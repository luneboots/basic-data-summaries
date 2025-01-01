import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

file_path = "sample-random-data.csv"
data = pd.read_csv(file_path)

first_100_rows = data.head(100)

country_summary = data.groupby("country").size().reset_index(name="sum-individuals")

data["date"] = pd.to_datetime(data["date"], format="%b %d, %Y")
data["join_month_year"] = data["date"].dt.strftime("%Y-%B")
monthly_yearly_counts = data["join_month_year"].value_counts().sort_index()

plt.figure(figsize=(10, 6))
plt.bar(monthly_yearly_counts.index, monthly_yearly_counts.values)
plt.title("Number of Individuals who Joined Month/Yr")
plt.xlabel("Month and Year")
plt.ylabel("Number of Joined Individuals")
plt.xticks(rotation=90)
plt.tight_layout()

image_file_path = "join_dates_histogram.png"
plt.savefig(image_file_path)
plt.close()

output_folder = "output_folder"
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

output_file_path = f"{output_folder}/output_data.xlsx"

with pd.ExcelWriter(output_file_path, engine="openpyxl") as writer:
    first_100_rows.to_excel(writer,sheet_name="print-rows", index=False)
    country_summary.to_excel(writer, sheet_name="country-summaries", index=False)

from openpyxl import load_workbook

book = load_workbook(output_file_path)
sheet_name = "join-dates-summaries"
sheet = book.create_sheet(sheet_name)

img = Image(image_file_path)
sheet.add_image(img, "B3")

book.save(output_file_path)

print(f"Dataset summaries have been saved to {output_file_path}")